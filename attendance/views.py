import csv
import io
import json
import uuid
from datetime import datetime
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, StreamingHttpResponse
from django.core.paginator import Paginator
from django.db.models import Q
from django.core.cache import cache
from .models import AttendanceRecord, UploadHistory
from .forms import AttendanceUploadForm, AttendanceEditForm, AttendanceFilterForm
from users.models import User, SupervisorAssignment

@login_required
def attendance_list(request):
    user = request.user
    form = AttendanceFilterForm(request.GET)
    
    # Base queryset based on user role
    if user.role == 'master':
        attendance_records = AttendanceRecord.objects.all()
    elif user.role == 'user1':
        attendance_records = AttendanceRecord.objects.filter(user__company_name=user.company_name)
    elif user.role == 'user2':
        assigned_employees = SupervisorAssignment.objects.filter(supervisor=user).values_list('employee', flat=True)
        attendance_records = AttendanceRecord.objects.filter(user__id__in=assigned_employees)
    elif user.role == 'user3':
        attendance_records = AttendanceRecord.objects.filter(user=user)
    else:
        attendance_records = AttendanceRecord.objects.none()
    
    # Apply filters
    if form.is_valid():
        if form.cleaned_data['start_date']:
            attendance_records = attendance_records.filter(date__gte=form.cleaned_data['start_date'])
        if form.cleaned_data['end_date']:
            attendance_records = attendance_records.filter(date__lte=form.cleaned_data['end_date'])
        if form.cleaned_data['status']:
            attendance_records = attendance_records.filter(status=form.cleaned_data['status'])
        if form.cleaned_data['employee']:
            attendance_records = attendance_records.filter(
                Q(user__username__icontains=form.cleaned_data['employee']) |
                Q(user__ep_number__icontains=form.cleaned_data['employee']) |
                Q(user__first_name__icontains=form.cleaned_data['employee']) |
                Q(user__last_name__icontains=form.cleaned_data['employee'])
            )
    
    attendance_records = attendance_records.select_related('user').order_by('-date')
    
    paginator = Paginator(attendance_records, 25)
    page = request.GET.get('page')
    attendance_records = paginator.get_page(page)
    
    return render(request, 'attendance/attendance_list.html', {
        'attendance_records': attendance_records,
        'form': form
    })

@login_required
def upload_attendance(request):
    if request.user.role not in ['master', 'user1']:
        messages.error(request, 'Permission denied.')
        return redirect('attendance_list')
    
    if request.method == 'POST':
        form = AttendanceUploadForm(request.POST, request.FILES)
        if form.is_valid():
            csv_file = request.FILES['csv_file']
            decoded_file = csv_file.read().decode('utf-8')
            io_string = io.StringIO(decoded_file)
            reader = csv.DictReader(io_string)
            
            # Generate session ID for progress tracking
            session_id = str(uuid.uuid4())
            request.session['upload_session_id'] = session_id
            
            created_count = 0
            updated_count = 0
            error_count = 0
            errors = []
            failed_rows = []
            
            # Store original headers for error file
            original_headers = reader.fieldnames
            
            # Count total rows for progress
            rows_list = list(reader)
            total_rows = len(rows_list)
            
            # Initialize progress
            progress_data = {
                'total': total_rows,
                'processed': 0,
                'success': 0,
                'errors': 0,
                'status': 'processing'
            }
            cache.set(f'upload_progress_{session_id}', progress_data, 300)  # 5 minutes
            
            for row_num, row in enumerate(rows_list, start=2):
                # Update progress
                progress_data['processed'] = row_num - 1
                cache.set(f'upload_progress_{session_id}', progress_data, 300)
                try:
                    ep_number = row.get('EP Number')
                    if not ep_number:
                        raise ValueError("EP Number is required")
                    
                    # Get or create user
                    user, created = User.objects.get_or_create(
                        ep_number=ep_number,
                        defaults={
                            'username': ep_number,
                            'first_name': row.get('Name', '').split()[0] if row.get('Name') else '',
                            'last_name': ' '.join(row.get('Name', '').split()[1:]) if row.get('Name') else '',
                            'company_name': row.get('Company Name'),
                            'plant': row.get('Plant'),
                            'department': row.get('Department'),
                            'trade': row.get('Trade'),
                            'skill': row.get('Skill'),
                            'shift': row.get('Shift'),
                            'role': 'user3'
                        }
                    )
                    
                    # Set password properly if user was just created
                    if created:
                        user.set_password(ep_number)
                        user.save()
                    
                    # Parse date
                    date_str = row.get('Date')
                    try:
                        date = datetime.strptime(date_str, '%d-%m-%Y').date()
                    except ValueError:
                        raise ValueError(f"Invalid date format: {date_str}. Expected DD-MM-YYYY")
                    
                    # Parse times (supports HH:MM and HH:MM (N) for next day)
                    def parse_time(time_str):
                        if not time_str or time_str.strip() == '':
                            return None
                        
                        time_str = time_str.strip()
                        
                        # Handle HH:MM (N) format - remove (N) indicator
                        if '(N)' in time_str or '(n)' in time_str:
                            time_str = time_str.replace('(N)', '').replace('(n)', '').strip()
                        
                        try:
                            return datetime.strptime(time_str, '%H:%M').time()
                        except ValueError:
                            raise ValueError(f"Invalid time format: {time_str}. Expected HH:MM or HH:MM (N)")
                    
                    # Parse hours and overtime (can be decimal or HH:MM format)
                    def parse_hours(hours_str):
                        if not hours_str or str(hours_str).strip() == '':
                            return 0
                        hours_str = str(hours_str).strip()
                        
                        # Check if it's in HH:MM format
                        if ':' in hours_str:
                            try:
                                parts = hours_str.split(':')
                                hours = int(parts[0])
                                minutes = int(parts[1]) if len(parts) > 1 else 0
                                return hours + (minutes / 60.0)
                            except (ValueError, IndexError):
                                return 0
                        else:
                            # Assume it's decimal format
                            try:
                                return float(hours_str)
                            except (ValueError, TypeError):
                                return 0
                    
                    # Create or update attendance record
                    attendance, created_record = AttendanceRecord.objects.update_or_create(
                        user=user,
                        date=date,
                        defaults={
                            'in1': parse_time(row.get('IN1')),
                            'out1': parse_time(row.get('OUT1')),
                            'in2': parse_time(row.get('IN2')),
                            'out2': parse_time(row.get('OUT2')),
                            'in3': parse_time(row.get('IN3')),
                            'out3': parse_time(row.get('OUT3')),
                            'hours_worked': parse_hours(row.get('Hours Worked')),
                            'overtime': parse_hours(row.get('Overtime')),
                            'status': row.get('Status', 'P'),
                            'shift': row.get('Shift')
                        }
                    )
                    
                    if created_record:
                        created_count += 1
                    else:
                        updated_count += 1
                    
                    # Update progress success count
                    progress_data['success'] = created_count + updated_count
                    cache.set(f'upload_progress_{session_id}', progress_data, 300)
                        
                except Exception as e:
                    error_count += 1
                    error_msg = str(e)
                    errors.append(f"Row {row_num}: {error_msg}")
                    
                    # Add failed row with error message for download
                    failed_row = dict(row)
                    failed_row['Error_Message'] = error_msg
                    failed_row['Row_Number'] = row_num
                    failed_rows.append(failed_row)
                    
                    # Update progress error count
                    progress_data['errors'] = error_count
                    cache.set(f'upload_progress_{session_id}', progress_data, 300)
            
            # Create error file if there are errors
            error_file_path = None
            if failed_rows:
                from django.core.files.base import ContentFile
                import tempfile
                import os
                
                # Create error CSV content
                error_output = io.StringIO()
                error_headers = ['Row_Number', 'Error_Message'] + original_headers
                error_writer = csv.DictWriter(error_output, fieldnames=error_headers)
                error_writer.writeheader()
                
                for failed_row in failed_rows:
                    error_writer.writerow(failed_row)
                
                # Save error file
                error_filename = f"errors_{csv_file.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
                error_content = ContentFile(error_output.getvalue().encode('utf-8'))
                
                # Create upload history with error file
                upload_history = UploadHistory.objects.create(
                    uploaded_by=request.user,
                    filename=csv_file.name,
                    total_rows=created_count + updated_count + error_count,
                    accepted_rows=created_count + updated_count,
                    rejected_rows=error_count
                )
                upload_history.error_file.save(error_filename, error_content)
                error_file_path = upload_history.error_file.url
            else:
                # Save upload history without error file
                UploadHistory.objects.create(
                    uploaded_by=request.user,
                    filename=csv_file.name,
                    total_rows=created_count + updated_count + error_count,
                    accepted_rows=created_count + updated_count,
                    rejected_rows=error_count
                )
            
            # Provide detailed feedback with statistics
            total_processed = created_count + updated_count
            total_rows = created_count + updated_count + error_count
            
            if error_count == 0:
                # All data uploaded successfully
                messages.success(request, f'✅ Upload Complete! Successfully processed all {total_rows} records')
                messages.info(request, f'📊 Statistics: {created_count} new records created, {updated_count} existing records updated')
            elif total_processed > 0:
                # Partial success
                messages.success(request, f'✅ Partial Success! {total_processed} out of {total_rows} records processed successfully')
                messages.info(request, f'📊 Statistics: {created_count} new, {updated_count} updated, {error_count} failed')
                messages.warning(request, f'⚠️ {error_count} records failed - Download error file to fix and re-upload')
                if error_file_path:
                    messages.error(request, f'📥 <a href="{error_file_path}" class="alert-link">Download Error File</a> to see failed records')
            else:
                # Complete failure
                messages.error(request, f'❌ Upload Failed! All {error_count} records could not be processed')
                messages.info(request, f'📊 Statistics: 0 successful, {error_count} failed out of {total_rows} total rows')
                if error_file_path:
                    messages.error(request, f'📥 <a href="{error_file_path}" class="alert-link">Download Error File</a> to fix issues and try again')
            
            return redirect('upload_attendance')
    else:
        form = AttendanceUploadForm()
    
    # Get recent upload history for display
    recent_uploads = UploadHistory.objects.filter(
        uploaded_by=request.user
    ).order_by('-upload_date')[:5]
    
    return render(request, 'attendance/upload_attendance.html', {
        'form': form,
        'recent_uploads': recent_uploads
    })

@login_required
def edit_attendance(request, pk):
    attendance = get_object_or_404(AttendanceRecord, pk=pk)
    
    # Check permissions
    if request.user.role == 'user3' and attendance.user != request.user:
        messages.error(request, 'Permission denied.')
        return redirect('attendance_list')
    elif request.user.role == 'user2':
        assigned_employees = SupervisorAssignment.objects.filter(supervisor=request.user).values_list('employee', flat=True)
        if attendance.user.id not in assigned_employees:
            messages.error(request, 'Permission denied.')
            return redirect('attendance_list')
    elif request.user.role == 'user1' and attendance.user.company_name != request.user.company_name:
        messages.error(request, 'Permission denied.')
        return redirect('attendance_list')
    
    if request.method == 'POST':
        form = AttendanceEditForm(request.POST, instance=attendance)
        if form.is_valid():
            form.save()
            messages.success(request, 'Attendance record updated successfully.')
            return redirect('attendance_list')
    else:
        form = AttendanceEditForm(instance=attendance)
    
    return render(request, 'attendance/edit_attendance.html', {'form': form, 'attendance': attendance})

@login_required
def export_attendance(request):
    if request.method == 'POST':
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        user = request.user
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        
        # Get filtered queryset
        if user.role == 'master':
            attendance_records = AttendanceRecord.objects.all()
        elif user.role == 'user1':
            attendance_records = AttendanceRecord.objects.filter(user__company_name=user.company_name)
        elif user.role == 'user2':
            assigned_employees = SupervisorAssignment.objects.filter(supervisor=user).values_list('employee', flat=True)
            attendance_records = AttendanceRecord.objects.filter(user__id__in=assigned_employees)
        else:
            attendance_records = AttendanceRecord.objects.none()
        
        # Apply date filters
        if start_date:
            attendance_records = attendance_records.filter(date__gte=start_date)
        if end_date:
            attendance_records = attendance_records.filter(date__lte=end_date)
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Report"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Headers
        headers = [
            'EP Number', 'Name', 'Company Name', 'Plant', 'Department', 'Trade', 'Skill', 'Shift',
            'Date', 'IN1', 'OUT1', 'IN2', 'OUT2', 'IN3', 'OUT3', 'Hours Worked', 'Overtime', 'Status',
            'Supervisor Remarks', 'Employee Remarks'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Data rows
        for row, record in enumerate(attendance_records.select_related('user'), 2):
            ws.cell(row=row, column=1, value=record.user.ep_number)
            ws.cell(row=row, column=2, value=f"{record.user.first_name} {record.user.last_name}".strip())
            ws.cell(row=row, column=3, value=record.user.company_name)
            ws.cell(row=row, column=4, value=record.user.plant)
            ws.cell(row=row, column=5, value=record.user.department)
            ws.cell(row=row, column=6, value=record.user.trade)
            ws.cell(row=row, column=7, value=record.user.skill)
            ws.cell(row=row, column=8, value=record.shift or record.user.shift)
            ws.cell(row=row, column=9, value=record.date.strftime('%d-%m-%Y'))
            ws.cell(row=row, column=10, value=record.in1.strftime('%H:%M') if record.in1 else '')
            ws.cell(row=row, column=11, value=record.out1.strftime('%H:%M') if record.out1 else '')
            ws.cell(row=row, column=12, value=record.in2.strftime('%H:%M') if record.in2 else '')
            ws.cell(row=row, column=13, value=record.out2.strftime('%H:%M') if record.out2 else '')
            ws.cell(row=row, column=14, value=record.in3.strftime('%H:%M') if record.in3 else '')
            ws.cell(row=row, column=15, value=record.out3.strftime('%H:%M') if record.out3 else '')
            ws.cell(row=row, column=16, value=record.get_hours_formatted())
            ws.cell(row=row, column=17, value=record.get_overtime_formatted())
            ws.cell(row=row, column=18, value=record.status)
            ws.cell(row=row, column=19, value=record.supervisor_remarks or '')
            ws.cell(row=row, column=20, value=record.employee_remarks or '')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"attendance_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    
    # Show export form
    return render(request, 'attendance/export_form.html')

@login_required
def download_attendance_template(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="attendance_template.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['EP Number', 'Name', 'Company Name', 'Plant', 'Department', 'Trade', 'Skill', 'Shift', 'Date', 'IN1', 'OUT1', 'IN2', 'OUT2', 'IN3', 'OUT3', 'Hours Worked', 'Overtime', 'Status'])
    writer.writerow(['EMP001', 'John Doe', 'ABC Company', 'Plant 1', 'Production', 'Welder', 'Skilled', 'Day', '01-01-2024', '08:00', '17:00', '', '', '', '', '8:00', '0:00', 'P'])
    writer.writerow(['EMP002', 'Jane Smith', 'ABC Company', 'Plant 1', 'Quality', 'Inspector', 'Semi-Skilled', 'Night', '01-01-2024', '22:30', '06:30 (N)', '', '', '', '', '8:00', '0:00', 'P'])
    writer.writerow(['EMP003', 'Mike Johnson', 'XYZ Corp', 'Plant 2', 'Maintenance', 'Technician', 'Skilled', 'Day', '01-01-2024', '', '', '', '', '', '', '0:00', '0:00', 'A'])
    
    return response

@login_required
def upload_history(request):
    if request.user.role not in ['master', 'user1']:
        messages.error(request, 'Permission denied.')
        return redirect('attendance_list')
    
    # Get upload history based on user role
    if request.user.role == 'master':
        uploads = UploadHistory.objects.all()
    else:
        uploads = UploadHistory.objects.filter(uploaded_by=request.user)
    
    uploads = uploads.order_by('-upload_date')
    
    paginator = Paginator(uploads, 20)
    page = request.GET.get('page')
    uploads = paginator.get_page(page)
    
    return render(request, 'attendance/upload_history.html', {'uploads': uploads})

@login_required
def upload_progress(request, session_id):
    """Return upload progress as JSON for real-time updates"""
    progress_data = cache.get(f'upload_progress_{session_id}', {
        'total': 0,
        'processed': 0,
        'success': 0,
        'errors': 0,
        'status': 'not_found'
    })
    
    return HttpResponse(
        json.dumps(progress_data),
        content_type='application/json'
    )